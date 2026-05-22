import re
from typing import Any, Dict

import openpyxl

from otp_relay.audit import audit
from otp_relay.config import logger
from otp_relay.state import users


USER_IMPORT_HEADER_ALIASES = {
    "token": "token",
    "user token": "token",
    "username": "token",
    "name": "name",
    "display name": "name",
    "email": "email",
    "email address": "email",
    "mail": "email",
    "test_env": "test_env",
    "test env": "test_env",
    "test environment": "test_env",
    "test_environment": "test_env",
    "testenv": "test_env",
    "prod_env": "prod_env",
    "prod env": "prod_env",
    "prod environment": "prod_env",
    "prod_environment": "prod_env",
    "production env": "prod_env",
    "production environment": "prod_env",
    "prodenv": "prod_env",
}


def _normalize_user_import_header(value: Any) -> str:
    clean = re.sub(r"\s+", " ", str(value or "").strip().lower().replace("-", " ").replace("_", " "))
    return USER_IMPORT_HEADER_ALIASES.get(clean, clean.replace(" ", "_"))


def _xlsx_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_users_from_excel(path: str, replace_existing: bool = True) -> int:
    """
    Reads users.xlsx. Expected columns (row 1 = headers):
      token    - 2 or 3 character unique string, e.g. AH or AHM
      name     - display name
      email    - company email address
    Optional columns:
      test_env - test environment assignment shown in the admin wizard
      prod_env - production environment assignment shown in the admin wizard
    Column names are case-insensitive. Spaces, dashes, and underscores are tolerated.
    Skipped rows are written to the audit log so IT can fix them.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    raw_headers = [
        _normalize_user_import_header(cell.value)
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    duplicate_headers = sorted({header for header in raw_headers if header and raw_headers.count(header) > 1})
    if duplicate_headers:
        raise ValueError(f"users.xlsx has duplicate column(s): {', '.join(duplicate_headers)}")

    required_headers = {"token", "name", "email"}
    missing_headers = sorted(required_headers - set(raw_headers))
    if missing_headers:
        raise ValueError(f"users.xlsx missing required column(s): {', '.join(missing_headers)}")

    loaded = 0
    skipped = 0
    seen_tokens: Dict[str, int] = {}
    imported_users: Dict[str, Dict[str, str]] = {}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None for value in row):
            continue

        row_dict = dict(zip(raw_headers, row))
        token = _xlsx_text(row_dict.get("token")).upper()
        name = _xlsx_text(row_dict.get("name"))
        email = _xlsx_text(row_dict.get("email"))
        test_env = _xlsx_text(row_dict.get("test_env"))
        prod_env = _xlsx_text(row_dict.get("prod_env"))

        if len(token) == 0:
            audit("import_skipped", detail=f"Row {row_num}: empty token - name={repr(name)} email={repr(email)}", status="warn")
            skipped += 1
            continue

        if not (2 <= len(token) <= 3):
            audit("import_skipped", token=token, detail=f"Row {row_num}: token must be 2 or 3 characters, got {len(token)} ({repr(token)})", status="warn")
            skipped += 1
            continue

        if not re.match(r"^[A-Z0-9]+$", token):
            audit("import_skipped", token=token, detail=f"Row {row_num}: token contains invalid characters ({repr(token)}) - only letters and digits allowed", status="warn")
            skipped += 1
            continue

        if not email:
            audit("import_skipped", token=token, detail=f"Row {row_num}: missing email address for {repr(name)}", status="warn")
            skipped += 1
            continue

        if "@" not in email:
            audit("import_skipped", token=token, detail=f"Row {row_num}: invalid email address {repr(email)}", status="warn")
            skipped += 1
            continue

        if token in seen_tokens:
            audit("import_skipped", token=token, detail=f"Row {row_num}: duplicate token - already defined at row {seen_tokens[token]}", status="warn")
            skipped += 1
            continue

        seen_tokens[token] = row_num
        imported_users[token] = {
            "token": token,
            "name": name,
            "email": email,
            "test_env": test_env,
            "prod_env": prod_env,
        }
        loaded += 1

    if replace_existing:
        users.clear()
        users.update(imported_users)

    logger.info("Loaded %s users from %s (%s rows skipped)", loaded, path, skipped)
    if skipped > 0:
        audit("import_complete", detail=f"{loaded} users loaded, {skipped} rows skipped - check import_skipped entries above", status="warn")
    else:
        audit("import_complete", detail=f"{loaded} users loaded, no issues")
    return loaded
